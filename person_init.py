import openpyxl

class Person:
    def __init__(self, first_name, last_name, class_year, parse_id, parents=[], children=[]):
        self.parse_id = parse_id
        self.first_name = first_name
        self.last_name = last_name
        self.class_year = class_year
        self.parents = parents
        self.children = children


def read_excel_data(filename):
    """
    Reads data from an Excel file and returns a list of Person objects
    """
    # Load the workbook and select the first worksheet
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    # Initialize a list to hold the Person objects
    people_list = []

    # Loop through each row in the worksheet until an empty row is reached
    row_num = 2  # start at row 2 to skip the header row
    while worksheet.cell(row=row_num, column=1).value is not None:
        # Get the data from the row and create a Person object
        id_num = worksheet.cell(row=row_num, column=1).value
        first_name = worksheet.cell(row=row_num, column=2).value
        last_name = worksheet.cell(row=row_num, column=3).value
        class_year = worksheet.cell(row=row_num, column=4).value

        parents = []
        for i in [5, 6, 7]:
            if worksheet.cell(row=row_num, column=i).value:
                parents.append(worksheet.cell(row=row_num, column=i).value)

        children = []
        for i in [8, 9, 10, 11, 12]:
            if worksheet.cell(row=row_num, column=i).value:
                children.append(worksheet.cell(row=row_num, column=i).value)
        person = Person(id_num, first_name, last_name, class_year, parents, children)

        # Add the Person object to the list
        people_list.append(person)

        # Move to the next row
        row_num += 1

    # Close the workbook and return the list of Person objects
    workbook.close()
    return people_list